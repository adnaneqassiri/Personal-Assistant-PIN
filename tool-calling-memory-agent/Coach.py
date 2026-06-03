import os
import sys
import uuid
import requests
import openpyxl
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from duckduckgo_search import DDGS
from db import get_users_col, get_conversations_col, find_or_create_user, get_user_history

load_dotenv()
USER_ID: str = os.getenv("USER_ID", "default_user")

from langchain_ollama import ChatOllama
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.tools import tool
from langchain_classic.agents import AgentExecutor, create_tool_calling_agent
from langchain_core.messages import HumanMessage, AIMessage
from langchain_core.callbacks.base import BaseCallbackHandler


# ─────────────────────────────────────────
# 1. CLASSES: PROFILE & STORAGE  (MongoDB-backed)
# ─────────────────────────────────────────

class UserProfile:
    """
    MongoDB-backed user profile.
    - Loads the user document ONCE at startup into an in-memory cache.
    - update() / delete() write through to MongoDB immediately via $set / $unset.
    - get_all() / to_prompt_string() read only from the cache — zero DB round-trips.
    """

    def __init__(self, user_id: str = USER_ID):
        self.user_id = user_id
        self._col = get_users_col()
        self._cache: dict = {}
        self._load_once()

    def _load_once(self):
        """Fetch the user document once; strip MongoDB internals."""
        doc = self._col.find_one({"user_id": self.user_id})
        if doc:
            self._cache = {k: v for k, v in doc.items() if k not in ("_id", "user_id")}

    def update(self, key: str, value: str):
        """Persist a single profile field via $set (upsert-safe)."""
        self._cache[key] = value
        self._col.update_one(
            {"user_id": self.user_id},
            {"$set": {key: value, "user_id": self.user_id}},
            upsert=True,
        )
        print(f"\n🔍 DEBUG: Profile saved → {key}: {value}")

    def delete(self, key: str):
        """Remove a single profile field via $unset."""
        if key in self._cache:
            del self._cache[key]
            self._col.update_one(
                {"user_id": self.user_id},
                {"$unset": {key: ""}},
            )

    def get_all(self) -> dict:
        return self._cache

    def to_prompt_string(self) -> str:
        if not self._cache:
            return "No information about the user yet."
        return "\n".join(f"- {k}: {v}" for k, v in self._cache.items())


class ConversationStorage:
    """
    MongoDB-backed conversation storage.
    - Creates one document per session in the 'conversations' collection.
    - save_message() uses $push to append messages and $set to update
      metadata — no full-document rewrites, no atexit flush needed.
    - print_history() re-reads the live document from MongoDB.
    """

    def __init__(self, user_id: str = USER_ID):
        self.user_id = user_id
        self.session_id = str(uuid.uuid4())[:8]
        self._col = get_conversations_col()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Insert the session document immediately so the session exists in DB
        self._col.insert_one({
            "session_id": self.session_id,
            "user_id": self.user_id,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "started_at": now,
            "messages": [],
            "total_messages": 0,
            "last_updated": now,
        })

    def save_message(self, role: str, message: str):
        """Append a message to the session document using $push."""
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "role": role,
            "message": message,
        }
        self._col.update_one(
            {"session_id": self.session_id},
            {
                "$push": {"messages": entry},
                "$inc":  {"total_messages": 1},
                "$set":  {"last_updated": entry["timestamp"]},
            },
        )

    def print_history(self):
        """Fetch and display the current session's messages from MongoDB."""
        doc = self._col.find_one({"session_id": self.session_id})
        if not doc or not doc.get("messages"):
            print("No messages yet.")
            return
        print(f"\n📅 Session: {self.session_id} — {doc['date']}")
        print("=" * 55)
        for msg in doc["messages"]:
            role = "You" if msg["role"] == "human" else "COACH"
            print(f"[{msg['timestamp']}] {role}: {msg['message']}\n")


# ─────────────────────────────────────────
# 2. HANDLERS & LOADERS
# ─────────────────────────────────────────

class StreamingHandler(BaseCallbackHandler):
    def on_llm_new_token(self, token: str, **kwargs):
        print(token, end="", flush=True)

    def on_tool_start(self, serialized, input_str, **kwargs):
        print(f"\n🔧 Using tool : {serialized.get('name', 'unknown')}")
        print(f"   ↳ Input     : {input_str}\n")

    def on_tool_end(self, output, **kwargs):
        preview = str(output)[:400]
        print(f"📥 Tool result : {preview}{'...' if len(str(output)) > 400 else ''}")
        print("\n🧠 Thinking ", end="", flush=True)


def load_document(file_path: str) -> str:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".txt":
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    elif ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        return "\n".join(page.extract_text() for page in reader.pages)
    else:
        raise ValueError("Only .pdf and .txt files are supported.")


# ─────────────────────────────────────────
# 3. TOOLS  (same logic; closures created once)
# ─────────────────────────────────────────

def make_profile_tools(profile: UserProfile):
    @tool
    def update_user_profile(key: str, value: str) -> str:
        """Updates user profile with info like name, job, hobbies, or goals."""
        profile.update(key, value)
        return f"✅ Got it! I'll remember that {key}: {value}"

    @tool
    def forget_user_info(key: str) -> str:
        """Removes a piece of information from the user profile."""
        profile.delete(key)
        return f"✅ Got it! I've forgotten: {key}"

    @tool
    def show_user_profile() -> str:
        """Shows everything COACH knows about the user."""
        data = profile.get_all()
        if not data:
            return "I don't know anything about you yet!"
        lines = "\n".join(f"  • {k}: {v}" for k, v in data.items())
        return f"Here's what I know about you:\n{lines}"

    return [update_user_profile, forget_user_info, show_user_profile]


@tool
def get_current_time() -> str:
    """Returns the current date and time."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

@tool
def search_web(query: str) -> str:
    """Searches the web for productivity tips or health advice."""
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=3))
            return (
                "\n\n".join(f"{i+1}. {r['title']}\n{r['body']}" for i, r in enumerate(results))
                if results else "No results found."
            )
    except Exception as e:
        return f"Search error: {e}"

@tool
def calculator(expression: str) -> str:
    """Evaluates a math expression."""
    try:
        return str(eval(expression, {"__builtins__": {}}))
    except Exception as e:
        return f"Error: {e}"

@tool
def get_weather(city: str) -> str:
    """Gets current weather for a city."""
    try:
        response = requests.get(f"https://wttr.in/{city}?format=3", timeout=5)
        return response.text
    except Exception as e:
        return f"Error: {e}"

@tool
def save_note(note: str) -> str:
    """Saves an important note or recommendation to a file."""
    with open("coach_notes.txt", "a", encoding="utf-8") as f:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        f.write(f"[{timestamp}] {note}\n")
    return "Note saved successfully!"

@tool
def rate_my_day(
    productive_hours: float, wasted_hours: float,
    sleep_hours: float, exercise_minutes: float
) -> str:
    """Calculates a productivity score out of 100."""
    score = 0
    score += min(productive_hours * 10, 40)
    score += max(0, 20 - wasted_hours * 5)
    score += 20 if 7 <= sleep_hours <= 9 else 10 if sleep_hours >= 6 else 0
    score += min(exercise_minutes / 3, 20)
    grade = (
        "🟢 Excellent" if score >= 80
        else "🟡 Good" if score >= 60
        else "🔴 Needs Work"
    )
    return f"Your day score: {score:.0f}/100 — {grade}"

@tool
def get_nutrition_advice(meal_description: str) -> str:
    """Searches for nutrition advice based on what the user ate."""
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(f"nutrition advice {meal_description}", max_results=2))
            return "\n".join(r["body"] for r in results)
    except Exception as e:
        return f"Error: {e}"

@tool
def analyze_sleep(bedtime: str, wake_time: str) -> str:
    """Analyzes sleep quality. Format: HH:MM (24h)."""
    fmt = "%H:%M"
    bed = datetime.strptime(bedtime, fmt)
    wake = datetime.strptime(wake_time, fmt)
    if wake < bed:
        wake += timedelta(days=1)
    duration = (wake - bed).seconds / 3600
    quality = (
        "✅ Optimal" if 7 <= duration <= 9
        else "⚠️ Slightly low" if duration >= 6
        else "🔴 Insufficient"
    )
    return f"Sleep duration: {duration:.1f} hours — {quality}"

@tool
def recommend_exercise(available_minutes: int, fitness_goal: str) -> str:
    """Recommends a workout based on available time and goal."""
    try:
        with DDGS() as ddgs:
            query = f"{available_minutes} minute workout for {fitness_goal}"
            results = list(ddgs.text(query, max_results=2))
            return "\n".join(r["body"] for r in results)
    except Exception as e:
        return f"Error: {e}"

@tool
def create_tomorrow_plan(tasks: str) -> str:
    """Saves a prioritized plan for tomorrow to a file."""
    task_list = [t.strip() for t in tasks.split(",")]
    with open("tomorrow_plan.txt", "w", encoding="utf-8") as f:
        f.write("📅 Tomorrow's Plan\n" + "=" * 30 + "\n")
        for i, task in enumerate(task_list, 1):
            f.write(f"{i}. {task}\n")
    return f"Plan saved with {len(task_list)} tasks."

@tool
def save_to_excel(
    date: str = "unknown", wake_time: str = "unknown", sleep_time: str = "unknown",
    productive_hours: float = 0.0, wasted_hours: float = 0.0,
    exercise_minutes: int = 0, meals_count: int = 0,
    productivity_stars: int = 3, mood: str = "unknown", notes: str = ""
) -> str:
    """Saves the day's summary to an Excel file."""
    file_path = "daily_log.xlsx"
    wb = openpyxl.load_workbook(file_path) if os.path.exists(file_path) else openpyxl.Workbook()
    ws = wb.active
    if not os.path.exists(file_path):
        ws.title = "Daily Log"
        headers = [
            "📅 Date", "⏰ Wake Up", "🌙 Sleep", "✅ Productive Hrs",
            "❌ Wasted Hrs", "🏃 Exercise (min)", "🍽️ Meals",
            "⭐ Productivity", "😊 Mood", "📝 Notes"
        ]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill("solid", fgColor="2E86AB")
            cell.font = Font(bold=True, color="FFFFFF")

    stars = "⭐" * productivity_stars + "☆" * (5 - productivity_stars)
    row_data = [
        date, wake_time, sleep_time, productive_hours, wasted_hours,
        exercise_minutes, meals_count, stars, mood, notes
    ]
    existing_row = next(
        (row[0].row for row in ws.iter_rows(min_row=2) if str(row[0].value) == date), None
    )
    if existing_row:
        for col, val in enumerate(row_data, 1):
            ws.cell(row=existing_row, column=col, value=val)
    else:
        ws.append(row_data)

    wb.save(file_path)
    return f"✅ Logged to Excel for {date}."


tools_list = [
    get_current_time, search_web, calculator, save_note,
    rate_my_day, analyze_sleep, recommend_exercise,
    create_tomorrow_plan, save_to_excel, get_weather, get_nutrition_advice
]


# ─────────────────────────────────────────
# 4. PROMPT  (built ONCE; profile string injected per-turn via {user_profile})
# ─────────────────────────────────────────

def build_prompt(log_content: str) -> ChatPromptTemplate:
    """
    OPTIMIZATION: log_content is static for the session, so the template
    is built once.  The live profile string is passed as a regular input
    key each turn — no template rebuild needed.
    """
    return ChatPromptTemplate.from_messages([
        ("system", f"""You are COACH, a personal productivity AI.

USER PROFILE:
{{user_profile}}

ACTIVITY LOG:
{log_content}

CRITICAL INSTRUCTION:
If the user mentions a name, job, goal, city, or any new fact about themselves,
you MUST call 'update_user_profile' BEFORE answering.
Do not just say you will remember it—actually use the tool.

STEP-BY-STEP PROCESS:
1. Scan user input for personal info.
2. If found, call update_user_profile.
3. Then, provide your coaching response.

Intent detection:
1. FULL ANALYSIS  → run sleep, rate, plan, excel tools.
2. SPECIFIC QUESTION → direct answer from log.
3. SAVE TO EXCEL → call save_to_excel immediately.
"""),
        ("placeholder", "{chat_history}"),
        ("human", "{input}"),
        ("placeholder", "{agent_scratchpad}"),
    ])


# ─────────────────────────────────────────
# 5. MAIN  — executor built ONCE, reused every turn
# ─────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser(description="COACH — Personal AI Productivity Assistant")
    parser.add_argument("log_file", nargs="?", help="Path to the daily log file (.txt or .pdf)")
    parser.add_argument("--user", default=USER_ID,
                        help="Unique username (e.g. zakaria). Two users with the same "
                             "display name must pick different usernames.")
    args = parser.parse_args()

    # Resolve username → stable UUID (creates account on first run)
    user_id  = find_or_create_user(args.user)
    username = args.user

    file_path   = args.log_file or input("📂 Path to log: ").strip()
    log_content = load_document(file_path)

    profile = UserProfile(user_id=user_id)
    storage = ConversationStorage(user_id=user_id)
    all_tools = tools_list + make_profile_tools(profile)
    handler = StreamingHandler()

    # ── Model recommendation ─────────────────────────────────────────────
    # CPU-only?  Prefer a 1-2 B model.  Ranked by speed on CPU (fastest first):
    #   qwen2.5:1.5b   ← fastest, surprisingly capable for short coaching tasks
    #   gemma2:2b
    #   phi3:mini
    #   mistral:7b-q4  ← if you have ≥16 GB RAM and don't mind waiting
    #
    # Change the model name below to whatever you have pulled in Ollama.
    # ─────────────────────────────────────────────────────────────────────
    MODEL = os.getenv("OLLAMA_MODEL", "gemma4:e2b")

    llm = ChatOllama(
        model=MODEL,
        streaming=True,
        callbacks=[handler],
        # These reduce CPU memory pressure and improve throughput:
        num_ctx=32768,      # increased context window to read whole context
        num_thread=max(1, os.cpu_count() - 1),  # leave 1 core for the OS
    )

    # OPTIMIZATION: build prompt and executor ONCE, not per turn
    prompt = build_prompt(log_content)
    agent = create_tool_calling_agent(llm, all_tools, prompt)
    executor = AgentExecutor(
        agent=agent,
        tools=all_tools,
        callbacks=[handler],
        verbose=False,
        max_iterations=6,   # guard against runaway tool loops
    )

    chat_history: list = []

    print("\n" + "=" * 55 + f"\n COACH ONLINE  [{MODEL}]\n" + "=" * 55)

    while True:
        try:
            user_input = input("\nYou: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n\n👋 Goodbye!")
            break

        if not user_input:
            continue
        if user_input.lower() == "quit":
            break
        if user_input.lower() == "history":
            storage.print_history()
            continue
        if user_input.lower() == "history all":
            sessions = get_user_history(user_id)
            if not sessions:
                print("📢 No conversation history found for this user.")
            else:
                print(f"\n📚 Full history for '{username}' — {len(sessions)} session(s)\n")
                for s in sessions:
                    print(f"📅 Session {s['session_id']}  |  {s['date']}  |  {s['total_messages']} messages")
                    print("-" * 55)
                    for msg in s.get("messages", []):
                        role = "You" if msg["role"] == "human" else "COACH"
                        print(f"  [{msg['timestamp']}] {role}: {msg['message']}")
                    print()
            continue

        try:
            storage.save_message("human", user_input)

            response = executor.invoke({
                "input": user_input,
                "chat_history": chat_history,
                "user_profile": profile.to_prompt_string(),   # ← injected here, no rebuild
            })
            answer = response["output"]

            print(f"\nCOACH: {answer}")
            storage.save_message("coach", answer)

            # Keep only the last 10 turns in memory to limit context size
            chat_history.extend([
                HumanMessage(content=user_input),
                AIMessage(content=answer),
            ])
            if len(chat_history) > 20:
                chat_history = chat_history[-20:]

        except Exception as e:
            print(f"\nCOACH Error: {e}")


if __name__ == "__main__":
    main()